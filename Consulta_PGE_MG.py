import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
import re
import openpyxl
import threading
import time
import base64
from datetime import datetime, timedelta
from tkinter.ttk import Progressbar
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import shutil  # Para mover arquivos

def configurar_chrome_options():
    chrome_options = Options()
    chrome_options.add_argument('--kiosk-printing')
    return chrome_options

def update_log(message):
    log_text.insert(tk.END, message + "\n")
    log_text.see(tk.END)

def update_progress(current, total):
    progress['value'] = (current / total) * 100
    root.update_idletasks()

def save_log_to_file(log_filename="process_log.txt"):
    with open(log_filename, "w") as log_file:
        log_file.write(log_text.get("1.0", tk.END))

def gerar_pdf_dinamico(driver, caminho_diretorio, nome_pdf):
    caminho_arquivo_pdf = os.path.join(caminho_diretorio, f"{nome_pdf}.pdf")
    result = driver.execute_cdp_cmd("Page.printToPDF", {
        "landscape": False,
        "displayHeaderFooter": False,
        "printBackground": True,
        "preferCSSPageSize": True
    })
    with open(caminho_arquivo_pdf, "wb") as file:
        file.write(base64.b64decode(result['data']))
    return caminho_arquivo_pdf

def validate_excel_file(excel_file):
    if not os.path.isfile(excel_file):
        messagebox.showerror("Erro", "O arquivo selecionado não é válido.")
        return False
    if not excel_file.endswith(('.xls', '.xlsx')):
        messagebox.showerror("Erro", "O arquivo precisa ser um arquivo Excel (.xls ou .xlsx).")
        return False
    return True

def rename_excel_file(excel_file, demanda_name):
    new_name = os.path.join(os.path.dirname(excel_file), f"{demanda_name}.xlsx")
    try:
        os.rename(excel_file, new_name)
        return new_name
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao renomear o arquivo: {e}")
        return excel_file

def gerar_relatorio_final(linhas_sucesso, linhas_erro, tempo_total):
    relatorio = f"Linhas processadas com sucesso: {linhas_sucesso}\n"
    relatorio += f"Linhas com erro: {linhas_erro}\n"
    relatorio += f"Tempo total: {tempo_total}\n"
    with open("relatorio_final.txt", "w") as rel_file:
        rel_file.write(relatorio)
    update_log(f"Relatório final salvo: {linhas_sucesso} sucesso(s), {linhas_erro} erro(s).")

def iniciar_processo(excel_file, start_row, caminho_diretorio, demanda_name):
    excel_file = rename_excel_file(excel_file, demanda_name)

    start_time = datetime.now()
    update_log(f"Processo iniciado às: {start_time.strftime('%d-%m-%Y %H:%M:%S')}")

    if not validate_excel_file(excel_file):
        return

    try:
        workbook = openpyxl.load_workbook(excel_file)
        # Sempre pegar a primeira aba:
        planilha = workbook.worksheets[0]
    except Exception as e:
        messagebox.showerror("Erro ao abrir a planilha", f"Erro: {e}")
        update_log(f"Erro ao abrir a planilha: {e}")
        return

    total_rows = planilha.max_row
    linha = int(start_row)

    saved = False
    if not saved:
        try:
            workbook.save(excel_file)
            saved = True
        except Exception as e:
            messagebox.showerror("Feche a Planilha", f"A planilha está aberta. Feche e clique no botão iniciar novamente")
            update_log(f"Erro ao salvar a planilha: {e}")
            return

    if not os.path.exists(caminho_diretorio):
        messagebox.showerror("Erro de Diretório", f"O diretório {caminho_diretorio} não existe.")
        update_log(f"Erro: O diretório {caminho_diretorio} não existe.")
        return
    elif not os.access(caminho_diretorio, os.W_OK):
        messagebox.showerror("Erro de Diretório", f"O diretório {caminho_diretorio} não é gravável.")
        update_log(f"Erro: O diretório {caminho_diretorio} não é gravável.")
        return

    chrome_options = configurar_chrome_options()
    driver_service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=driver_service, options=chrome_options)

    link = 'http://receitaonline.fazenda.mg.gov.br/rol/dae/'
    driver.get(link)

    select_element = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[1]/tbody/tr[1]/td[2]/select"))
    )
    select = Select(select_element)

    try:
        select.select_by_visible_text('CNPJ')
    except Exception:
        pass

    linhas_sucesso = 0
    linhas_erro = 0
    pdf_paths = []

    # Variáveis para cálculo de tempo e atualizações na GUI
    processed_count = 0
    total_to_process = total_rows - int(start_row) + 1
    # Atualização inicial do status
    update_status_label(processed_count, total_to_process, linhas_sucesso, linhas_erro, start_time)

    while linha <= total_rows:
        coluna_5 = planilha.cell(row=linha, column=5).value

        if coluna_5 == "OK":
            linha += 1
            processed_count += 1
            # Atualiza status após cada linha
            update_status_label(processed_count, total_to_process, linhas_sucesso, linhas_erro, start_time)
            continue

        try:
            Num_CDA = planilha.cell(row=linha, column=3).value
            Num_CNPJ = planilha.cell(row=linha, column=2).value

            if Num_CNPJ is not None:
                Num_CNPJ = re.sub(r'\D', '', str(Num_CNPJ))

            if Num_CDA is not None:
                Num_CDA = re.sub(r'\D', '', str(Num_CDA))

            if Num_CDA is None:
                break

            Campo_CNPJ = driver.find_element(By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[1]/tbody/tr[2]/td[2]/input")
            Campo_CNPJ.clear()
            Campo_CNPJ.send_keys(Num_CNPJ)

            Campo_CDA = driver.find_element(By.XPATH, "//*[@id='id_numero_daf']")
            Campo_CDA.clear()
            Campo_CDA.send_keys(Num_CDA)

            btn_Consultar = driver.find_element(By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[2]/tbody/tr/td/input[1]")
            btn_Consultar.click()

            try:
                Valor_Total_Elemento = driver.find_element(By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[9]/td[2]/b/font")
                Valor_Total = Valor_Total_Elemento.text
            except NoSuchElementException:
                Valor_Total = "Não permite pagamento"

            planilha.cell(row=linha, column=4, value=Valor_Total)
            planilha.cell(row=linha, column=5, value="OK")

            workbook.save(excel_file)

            pdf_path = gerar_pdf_dinamico(driver, caminho_diretorio, str(Num_CDA))
            pdf_paths.append(pdf_path)
            update_log(f"Processada linha {linha}: CDA {Num_CDA}")
            linhas_sucesso += 1

        except (NoSuchElementException, TimeoutException) as e:
            update_log(f"Erro na linha {linha}: {e}")
            linhas_erro += 1
            planilha.cell(row=linha, column=4, value="Não permite pagamento")
            planilha.cell(row=linha, column=5, value="OK")
            workbook.save(excel_file)

        finally:
            linha += 1
            processed_count += 1
            update_progress(processed_count, total_to_process)
            update_status_label(processed_count, total_to_process, linhas_sucesso, linhas_erro, start_time)
            driver.back()

    end_time = datetime.now()
    total_time = end_time - start_time
    formatted_time = str(total_time).split(".")[0]

    driver.quit()
    save_log_to_file()
    gerar_relatorio_final(linhas_sucesso, linhas_erro, formatted_time)

    download_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    pasta_demanda = os.path.join(download_folder, demanda_name)
    if not os.path.exists(pasta_demanda):
        os.makedirs(pasta_demanda)
        update_log(f"Pasta '{demanda_name}' criada nos Downloads.")

    excel_novo_caminho = os.path.join(pasta_demanda, os.path.basename(excel_file))
    try:
        shutil.move(excel_file, excel_novo_caminho)
        update_log(f"Arquivo Excel movido para: {excel_novo_caminho}")
    except Exception as e:
        update_log(f"Erro ao mover o arquivo Excel: {e}")

    for pdf_path in pdf_paths:
        try:
            shutil.move(pdf_path, pasta_demanda)
            update_log(f"PDF {os.path.basename(pdf_path)} movido para: {pasta_demanda}")
        except Exception as e:
            update_log(f"Erro ao mover o PDF {os.path.basename(pdf_path)}: {e}")

    messagebox.showinfo("Processo concluído", f"Todas as CDAs estão preenchidas com sucesso!\n\nTempo total: {formatted_time}")
    update_log(f"Processo finalizado às: {end_time.strftime('%d-%m-%Y %H:%M:%S')}\nTempo total: {formatted_time}")

def update_status_label(processed_count, total_count, success_count, error_count, start_time):
    elapsed = datetime.now() - start_time
    elapsed_seconds = elapsed.total_seconds()
    if processed_count > 0:
        avg_time_per_line = elapsed_seconds / processed_count
        remaining = total_count - processed_count
        eta_seconds = avg_time_per_line * remaining
        eta = datetime.now() + timedelta(seconds=eta_seconds)
        eta_str = eta.strftime('%H:%M:%S')
    else:
        eta_str = "--:--:--"

    status_msg = (f"Processadas: {processed_count}/{total_count} | " f"Sucesso: {success_count} | Erros: {error_count} | " f"Tempo decorrido: {str(elapsed).split('.')[0]} | ETA: {eta_str}")
    stats_label.config(text=status_msg)
    root.update_idletasks()

def browse_file():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, filename)

def browse_directory():
    directory = filedialog.askdirectory()
    directory_entry.delete(0, tk.END)
    directory_entry.insert(0, directory)

def on_start():
    excel_file = excel_file_entry.get()
    start_row = 2
    caminho_diretorio = directory_entry.get()
    demanda_name = demanda_entry.get()

    if not excel_file or not start_row or not caminho_diretorio or not demanda_name:
        messagebox.showwarning("Erro de Entrada", "Por favor, preencha todos os campos.")
        return

    threading.Thread(target=iniciar_processo, args=(excel_file, start_row, caminho_diretorio, demanda_name), daemon=True).start()

root = tk.Tk()
root.title("Automação PGE-MG")

tk.Label(root, text="Aponte o arquivo Excel:").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
excel_file_entry = tk.Entry(root, width=100)
excel_file_entry.grid(row=0, column=1, padx=10, pady=5)
tk.Button(root, text="Buscar", command=browse_file).grid(row=0, column=2, padx=10, pady=5)

tk.Label(root, text="Aponte o diretório para Salvar PDFs:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
directory_entry = tk.Entry(root, width=100)
directory_entry.grid(row=1, column=1, padx=10, pady=5)
tk.Button(root, text="Buscar", command=browse_directory).grid(row=1, column=2, padx=10, pady=5)

tk.Label(root, text="Nome da Demanda:").grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
demanda_entry = tk.Entry(root, width=50)
demanda_entry.grid(row=2, column=1, padx=10, pady=5)

tk.Button(root, text="Iniciar Processo", command=on_start).grid(row=3, column=0, columnspan=3, pady=10)

# Barra de status com informações do andamento
stats_label = tk.Label(root, text="Processadas: 0/0 | Sucesso: 0 | Erros: 0 | Tempo decorrido: 00:00:00 | ETA: --:--:--")
stats_label.grid(row=4, column=0, columnspan=3, pady=5)

progress = Progressbar(root, orient=tk.HORIZONTAL, length=400, mode='determinate')
progress.grid(row=5, column=0, columnspan=3, pady=10)

log_text = scrolledtext.ScrolledText(root, height=10, width=120)
log_text.grid(row=6, column=0, columnspan=3, padx=10, pady=10)

root.mainloop()
